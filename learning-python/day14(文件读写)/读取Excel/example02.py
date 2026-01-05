"""
example02 - 写一个Excel文件

Author: lsy
Date: 2026/1/5
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# 创建一个新的工作簿
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "员工信息表"

# 设置表头
headers = ["工号", "姓名", "部门", "职位", "入职日期", "基本工资", "绩效评分"]
for col, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col)
    cell.value = header
    # 设置表头样式
    cell.font = Font(name='微软雅黑', size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 模拟员工数据
employees = [
    ["001", "张三", "技术部", "软件工程师", "2023-01-15", 12000, 4.5],
    ["002", "李四", "产品部", "产品经理", "2022-06-20", 15000, 4.2],
    ["003", "王五", "市场部", "市场专员", "2023-03-10", 8000, 4.8],
    ["004", "赵六", "技术部", "前端工程师", "2023-07-01", 11000, 4.6],
    ["005", "钱七", "人力资源部", "HR专员", "2022-11-05", 9000, 4.3],
    ["006", "孙八", "财务部", "会计", "2021-09-15", 10000, 4.7],
    ["007", "周九", "技术部", "后端工程师", "2023-02-28", 13000, 4.4],
    ["008", "吴十", "市场部", "销售经理", "2020-12-01", 18000, 4.9],
]

# 填入数据
for row_idx, employee in enumerate(employees, 2):
    for col_idx, value in enumerate(employee, 1):
        cell = sheet.cell(row=row_idx, column=col_idx)
        cell.value = value
        # 设置数据样式
        cell.font = Font(name='微软雅黑', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # 为绩效评分设置条件格式颜色
        if col_idx == 7:  # 绩评分数列
            if value >= 4.5:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif value >= 4.0:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# 设置列宽
column_widths = [10, 10, 12, 12, 15, 12, 12]
for col, width in enumerate(column_widths, 1):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# 添加底部统计信息
sheet.cell(row=11, column=1, value="统计信息")
sheet.cell(row=11, column=1).font = Font(bold=True)
sheet.cell(row=12, column=1, value=f"员工总数: {len(employees)}")
sheet.cell(row=13, column=1, value=f"平均工资: {sum(emp[5] for emp in employees) / len(employees):.2f}")
sheet.cell(row=14, column=1, value=f"平均绩效: {sum(emp[6] for emp in employees) / len(employees):.2f}")

# 保存文件
file_path = "../resources/员工信息表.xlsx"
wb.save(file_path)
print(f"Excel文件已成功创建: {file_path}")

