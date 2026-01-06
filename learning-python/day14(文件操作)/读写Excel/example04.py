"""
example04 - 

Author: lsy
Date: 2026/1/5
"""
import openpyxl

wb = openpyxl.load_workbook('../resources/员工信息表.xlsx')
# print(wb.sheetnames)
sheet = wb['员工信息表'] # :type:
# print(sheet.dimensions)
# print(sheet.max_row, sheet.max_column)
#
for row in range(1, sheet.max_row - 1):
    for col in range(1,sheet.max_column + 1):
        print(sheet.cell(row,col).value,end=" ")
    print()


# print(sheet['B2'].value)

# for row in range(1, sheet.max_row-1):
#     for col in 'ABCDEFG':
#         print(sheet[f'{col}{row}'].value)

