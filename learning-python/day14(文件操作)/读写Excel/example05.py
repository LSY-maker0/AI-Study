"""
example05 - 写Excel文件

Author: lsy
Date: 2026/1/5
"""
import openpyxl
from datetime import datetime

# 1. 创建一个新的工作簿 (Workbook)
wb = openpyxl.Workbook()

# 2. 获取当前活跃的工作表
sheet = wb.active
sheet.title = "销售记录"  # 修改工作表名称

# sheet.create_sheet('销售记录')

# 3. 准备要写入的数据 (模拟一个销售列表)
headers = ["订单号", "产品名称", "单价", "数量", "总价", "销售日期"]
data_rows = [
    [1001, "机械键盘", 399, 2, 798, "2023-10-01"],
    [1002, "无线鼠标", 129, 5, 645, "2023-10-02"],
    [1003, "显示器", 1599, 1, 1599, "2023-10-03"],
    [1004, "USB转接器", 59, 10, 590, "2023-10-05"],
]

# 写入头部
for col_num, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_num, value=header)

# 写数据行
for row_num, row_data in enumerate(data_rows, start=2):
    for col_num, cell_value in enumerate(row_data, start=1):
        sheet.cell(row=row_num, column=col_num, value=cell_value)


# 保存到文件
filename = '../resources/销售记录.xlsx'
wb.save(filename)
print(f"Excel文件 '{filename}' 生成成功！")

